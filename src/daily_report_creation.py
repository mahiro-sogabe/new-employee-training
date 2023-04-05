import openpyxl

def create_daily_report():
    # Excelファイルを読み込みます。存在しない場合は新規作成します。
    try:
        wb = openpyxl.load_workbook('daily_report.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.active.title = 'Daily Report'
    sheet = wb.active

    # 氏名、報告日、作業内容、反省点を入力します。
    name = input('氏名を入力してください：')
    report_date = input('報告日を入力してください（例：2022/01/01）：')
    task = input('作業内容を入力してください：')
    reflection = input('反省点を入力してください：')

    # データをExcelファイルに書き込みます。
    if sheet.max_row == 1:
        sheet['A1'] = '氏名'
        sheet['B1'] = '報告日'
        sheet['C1'] = '作業内容'
        sheet['D1'] = '反省点'
    row = (name, report_date, task, reflection)
    sheet.append(row)

    # Excelファイルを保存します。
    file_name = f'日報_{name}_{report_date.replace("/", "")}.xlsx'
    wb.save(f'./out/{file_name}')
    print(f'{file_name}を作成しました。')

if __name__ == '__main__':
    create_daily_report()